#https://www.geeksforgeeks.org/python-writing-excel-file-using-openpyxl-module/ < tutorial
#https://openpyxl.readthedocs.io/en/stable/ < documentation
from bs4 import BeautifulSoup
import requests
#for webscraping
import openpyxl
from openpyxl.styles import Font
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
#for putting into excel

wb = openpyxl.Workbook()
sheet = wb.active
  

columnCounter = 1 #First column to insert to
rowCounter = 2 #First row to insert to


sheet["A1"] = "Title"
sheet["A1"].font = Font(bold=True)
sheet["B1"] = "Price"
sheet["B1"].font = Font(bold=True)
sheet["C1"] = "Rating"
sheet["C1"].font = Font(bold=True)

cell = sheet.cell(row = rowCounter, column = columnCounter)

firstColumn = 1
aColumn = str(chr(64 + firstColumn))
sheet.column_dimensions[aColumn].width = 65

thirdColumn = 3
cColumn = str(chr(64 + thirdColumn))
sheet.column_dimensions[cColumn].width = 20

URL = ["https://www.newegg.com/Desktop-Graphics-Cards/SubCategory/ID-48?Tid=7709", "https://www.newegg.com/Desktop-Graphics-Cards/SubCategory/ID-48/Page-2?Tid=7709", "https://www.newegg.com/Desktop-Graphics-Cards/SubCategory/ID-48/Page-3?Tid=7709"]
gpu = input("Enter desired GPU or leave empty for any ")
maxPrice = input("Enter max price ")

for url in URL:    
    page = requests.get(url)
    soup = BeautifulSoup(page.content, "html.parser")
    results = soup.find(class_="list-wrap")
    container = soup.find_all("div", class_="item-container")
    print()

    for containerElement in container:
        brandingElement = containerElement.find("div", class_="item-branding")
        titleElement = containerElement.find("a", class_="item-title")
        title = str(titleElement.contents)
        priceElement = containerElement.find("li", class_="price-current")
        price = str(priceElement.text)
        ratingElement = containerElement.find("i", class_="rating")
        
  
        for char in title:
            title = title.replace("[", "")
            title = title.replace("]","")
            title = title.replace("'", "")
        for char in price:
            price = price.replace("\xa0","")
            price = price.replace("–", "")
            price = price.replace("$", "")
            price = price.replace(",", "")
            price = price.split("(")[0]
        
        
        price = int(float(price))
        maxPrice = int(float(maxPrice))
        if gpu in title and price <= maxPrice:       
            print(title.split("GDDR")[0])
            cell = sheet.cell(row = rowCounter, column = columnCounter)
            cell.value = title.split("GDDR")[0]
            columnCounter += 1
            print("$" + str(price))
            cell = sheet.cell(row = rowCounter, column = columnCounter)
            cell.value = "$" + str(price)
            columnCounter += 1
            if ratingElement != None:
                rating = brandingElement.find("i", {"class":"rating"})["aria-label"]
                print(rating.capitalize())
                cell = sheet.cell(row = rowCounter, column = columnCounter)
                cell.value = rating.capitalize()
                columnCounter += 1
            else:
                print("No rating")
                cell = sheet.cell(row = rowCounter, column = columnCounter)
                cell.value = "No rating"
                columnCounter += 1
            print()
            rowCounter += 1
            columnCounter = 1    
        elif gpu == "" and price <= maxPrice:
            print(title.split("GDDR")[0])
            cell.value = title.split("GDDR")[0]
            cell = sheet.cell(row = rowCounter, column = columnCounter)
            columnCounter += 1
            print("$" + str(price))
            cell.value = "$" + str(price)
            cell = sheet.cell(row = rowCounter, column = columnCounter)
            columnCounter += 1
            if ratingElement != None:
                rating = brandingElement.find("i", {"class":"rating"})["aria-label"]
                print(rating.capitalize())
                cell.value = rating.capitalize()
                cell = sheet.cell(row = rowCounter, column = columnCounter)
                columnCounter += 1
            else:
                print("No rating")
                cell.value = "No rating"
                cell = sheet.cell(row = rowCounter, column = columnCounter)
                columnCounter += 1    
            print()
            rowCounter += 1
            columnCounter = 1
          
wb.save("C:\\Utilities\\Comp Sci Topics Stuff\\data.xlsx")